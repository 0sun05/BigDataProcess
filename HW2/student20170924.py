#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

total = {}
i = 1
for row in ws:
	if i != 1:
		ws.cell(row = i, column = 7).value = \
		ws.cell(row = i, column = 3).value * 0.3 + \
		ws.cell(row = i, column = 4).value * 0.35 + \
		ws.cell(row = i, column = 5).value * 0.34 + \
		ws.cell(row = i, column = 6).value
		total[i] = ws.cell(row = i, column = 7).value
	i += 1

sort_total = sorted(total.items(), key = lambda x : x[1], reverse=True)

for i in range(len(sort_total)):
	if i + 1 <= len(sort_total) / 10 * 3:
		if i + 1 <= 0 + len(sort_total) / 10 * (3 - 0) / 2:
			ws.cell(row = sort_total[i][0], column = 8).value ='A+'
		else:
			ws.cell(row = sort_total[i][0], column = 8).value ='A0'
	elif i + 1 <= len(sort_total) / 10 * 7:
		if i + 1 <= len(sort_total) / 10 * 3 +\
		len(sort_total) / 10 * (7 - 3) / 2:
			ws.cell(row = sort_total[i][0], column = 8).value ='B+'
		else:
			ws.cell(row = sort_total[i][0], column = 8).value ='B0'
	else:
		if i + 1 <= len(sort_total) / 10 * 7 +\
		len(sort_total) / 10 * (10 - 7) / 2:
			ws.cell(row = sort_total[i][0], column = 8).value ='C+'
		else:
			ws.cell(row = sort_total[i][0], column = 8).value ='C0'

for i in range(len(sort_total) - 1, 0, -1):
	if len(sort_total) != 1:
		if ws.cell(row = sort_total[i][0], column = 7).value ==\
		ws.cell(row = sort_total[i - 1][0], column = 7).value:
			ws.cell(row = sort_total[i-1][0], column = 8).value ==\
			ws.cell(row = sort_total[i][0], column = 8).value

wb.save("student.xlsx")
